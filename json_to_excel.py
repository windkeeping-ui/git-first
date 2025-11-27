#!/usr/bin/env python3
"""
Convert JSON news data to Excel format with date in filename
"""
import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd


def convert_json_to_excel(json_file: str, output_file: str = None):
    """Convert JSON file to Excel with styling and date in filename"""
    
    # Load JSON
    json_path = Path(json_file)
    if not json_path.exists():
        print(f"Error: {json_file} not found")
        sys.exit(1)
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print(f"Loaded {len(data)} articles from {json_file}")
    
    # Generate output filename with date if not specified
    if output_file is None:
        date_str = datetime.now().strftime("%Y%m%d")
        base_name = json_path.stem  # Remove .json extension
        output_file = f"{base_name}_{date_str}.xlsx"
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to Excel
    df.to_excel(output_file, sheet_name='Articles', index=False)
    print(f"Exported to {output_file}")
    
    # Apply styling
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Header styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header style
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Apply data styling
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # url
    ws.column_dimensions['B'].width = 50  # title
    ws.column_dimensions['C'].width = 60  # description
    ws.column_dimensions['D'].width = 25  # published
    if len(df.columns) > 4:
        ws.column_dimensions['E'].width = 80  # content
    if len(df.columns) > 5:
        ws.column_dimensions['F'].width = 80  # summary
    if len(df.columns) > 6:
        ws.column_dimensions['G'].width = 25  # scraped_at
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)
    print(f"Styling applied")
    
    # File stats
    import os
    file_size = os.path.getsize(output_file) / 1024
    print(f"File size: {file_size:.1f} KB")
    print(f"\nExcel file created: {output_file}")
    
    return output_file


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python json_to_excel.py <json_file> [output_file]")
        print("  If output_file is not specified, it will be generated with current date")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    convert_json_to_excel(json_file, output_file)
